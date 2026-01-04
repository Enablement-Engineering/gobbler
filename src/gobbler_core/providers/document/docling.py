"""Docling document conversion provider.

This provider uses the Docling Docker service for document conversion
to markdown with optional OCR support.
"""

from __future__ import annotations

import logging
import tempfile
from pathlib import Path
from typing import Any

import aiofiles

from gobbler_core.providers.document.base import DocumentProvider, DocumentResult
from gobbler_core.providers.registry import ProviderRegistry
from gobbler_core.utils.http_client import RetryableHTTPClient

logger = logging.getLogger(__name__)

# Supported document formats (including legacy .xls which we convert to .xlsx)
SUPPORTED_EXTENSIONS: set[str] = {".pdf", ".docx", ".pptx", ".xlsx", ".xls"}


class DoclingProvider(DocumentProvider):
    """Document conversion provider using Docling Docker service.

    Converts PDF, DOCX, PPTX, and XLSX documents to markdown using
    the Docling service running in Docker. Supports OCR for scanned
    documents.

    Attributes:
        service_url: URL of the Docling service

    Example:
        provider = DoclingProvider(service_url="http://localhost:5001")
        result = await provider.convert(Path("document.pdf"), ocr=True)
        print(result.markdown)
    """

    def __init__(
        self,
        service_url: str = "http://localhost:5001",
        timeout: float = 120.0,
    ) -> None:
        """Initialize the Docling provider.

        Args:
            service_url: URL of the Docling service
            timeout: Request timeout in seconds
        """
        self.service_url = service_url.rstrip("/")
        self.timeout = timeout

    @property
    def name(self) -> str:
        """Return provider name."""
        return "docling"

    def _convert_xls_to_xlsx(self, xls_path: Path) -> Path:
        """Convert legacy .xls file to .xlsx format.

        Uses xlrd to read the .xls file and openpyxl to write .xlsx.

        Args:
            xls_path: Path to the .xls file

        Returns:
            Path to temporary .xlsx file (caller must delete)

        Raises:
            RuntimeError: If conversion fails
        """
        try:
            import xlrd
            from openpyxl import Workbook
        except ImportError as e:
            msg = (
                "xlrd and openpyxl are required for .xls file support. "
                "Install with: pip install xlrd openpyxl"
            )
            raise RuntimeError(msg) from e

        try:
            # Read the .xls file
            xls_book = xlrd.open_workbook(str(xls_path))

            # Create a new .xlsx workbook
            xlsx_book = Workbook()
            # Remove the default sheet created by Workbook()
            default_sheet = xlsx_book.active
            if default_sheet is not None:
                xlsx_book.remove(default_sheet)

            # Copy each sheet
            for sheet_idx in range(xls_book.nsheets):
                xls_sheet = xls_book.sheet_by_index(sheet_idx)
                xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell_value = xls_sheet.cell_value(row_idx, col_idx)
                        xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

            # Write to temporary file
            temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
            import os

            os.close(temp_fd)
            xlsx_book.save(temp_path)

            logger.debug("Converted %s to temporary %s", xls_path, temp_path)
            return Path(temp_path)

        except Exception as e:
            msg = f"Failed to convert .xls to .xlsx: {e}"
            raise RuntimeError(msg) from e

    async def convert(
        self,
        file_path: Path,
        ocr: bool = True,
        **options: Any,  # noqa: ARG002  # Reserved for future provider options
    ) -> DocumentResult:
        """Convert document to markdown using Docling service.

        Args:
            file_path: Path to document file
            ocr: Enable OCR for scanned documents
            **options: Additional options (currently unused)

        Returns:
            DocumentResult with markdown content and metadata

        Raises:
            FileNotFoundError: If file_path doesn't exist
            ValueError: If file format is not supported
            RuntimeError: If conversion fails or service unavailable
        """
        # Validate file exists
        if not file_path.exists():
            msg = f"Document file not found: {file_path}"
            raise FileNotFoundError(msg)

        # Validate format
        ext = file_path.suffix.lower()
        if not self.supports_format(ext):
            msg = f"Unsupported format: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
            raise ValueError(msg)

        # Handle legacy .xls files by converting to .xlsx first
        temp_xlsx_path: Path | None = None
        original_path = file_path

        if ext == ".xls":
            logger.info("Converting legacy .xls file to .xlsx for processing")
            temp_xlsx_path = self._convert_xls_to_xlsx(file_path)
            file_path = temp_xlsx_path

        try:
            # Read file
            try:
                async with aiofiles.open(file_path, "rb") as f:
                    file_data = await f.read()
            except Exception as e:
                msg = f"Failed to read document file: {e}"
                raise RuntimeError(msg) from e

            # Make request to Docling service
            try:
                async with RetryableHTTPClient(timeout=self.timeout) as client:
                    files = {"files": (file_path.name, file_data)}
                    data = {
                        "to_formats": "md",
                        "do_ocr": str(ocr).lower(),
                    }

                    response = await client.post(
                        f"{self.service_url}/v1/convert/file",
                        files=files,
                        data=data,
                    )
                    response.raise_for_status()
                    result = response.json()

            except Exception as e:
                error_str = str(e)
                error_type = type(e).__name__

                if "ConnectError" in error_type or "Connection" in error_str:
                    msg = (
                        "Docling service unavailable. The service may not be running. "
                        "Start with: docker-compose up -d docling"
                    )
                    raise RuntimeError(msg) from e

                msg = f"Document conversion failed: {e}"
                raise RuntimeError(msg) from e

            # Process response (use original path for metadata)
            return self._process_response(result, original_path)

        finally:
            # Clean up temporary .xlsx file if we created one
            if temp_xlsx_path is not None and temp_xlsx_path.exists():
                try:
                    temp_xlsx_path.unlink()
                    logger.debug("Cleaned up temporary file: %s", temp_xlsx_path)
                except OSError as e:
                    logger.warning("Failed to clean up temporary file %s: %s", temp_xlsx_path, e)

    def _process_response(self, result: dict[str, Any], file_path: Path) -> DocumentResult:
        """Process the Docling API response.

        Args:
            result: JSON response from Docling
            file_path: Original file path for metadata

        Returns:
            DocumentResult

        Raises:
            RuntimeError: If conversion failed
        """
        status = result.get("status")

        if status == "failure":
            errors = result.get("errors", ["Unknown error"])
            msg = f"Document conversion failed: {'; '.join(errors)}"
            raise RuntimeError(msg)

        if status == "skipped":
            msg = (
                "Document conversion was skipped. The file may be corrupted or "
                "use an unsupported format variation."
            )
            raise RuntimeError(msg)

        document_data = result.get("document", {})
        markdown_content = document_data.get("md_content", "")

        if not markdown_content:
            msg = (
                "Failed to extract markdown from document. The document may be "
                "corrupted or password-protected."
            )
            raise RuntimeError(msg)

        # Estimate page count
        word_count = len(markdown_content.split())
        pages = result.get("pages", 0) if "pages" in result else max(1, word_count // 300)

        return DocumentResult(
            markdown=markdown_content,
            pages=pages,
            metadata={
                "file_path": str(file_path),
                "format": file_path.suffix.lower(),
                "service": "docling",
            },
        )

    def supports_format(self, file_extension: str) -> bool:
        """Check if file format is supported.

        Args:
            file_extension: File extension including dot (e.g., ".pdf")

        Returns:
            True if format is supported
        """
        return file_extension.lower() in SUPPORTED_EXTENSIONS


# Register provider with the registry
ProviderRegistry.register("document", "docling", DoclingProvider)
