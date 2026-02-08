"""Document conversion module with pluggable provider support.

This module provides document conversion capabilities with support for
pluggable conversion providers. The default provider uses the Docling
Docker service.

Example:
    # Using default provider
    markdown, metadata = await convert_document_to_markdown("document.pdf")

    # Using a specific provider
    from gobbler_core.providers.document import DoclingProvider
    provider = DoclingProvider(service_url="http://localhost:5001")
    markdown, metadata = await convert_document_to_markdown("document.pdf", provider=provider)
"""

from __future__ import annotations

import logging
import time
from collections.abc import Callable
from pathlib import Path
from typing import TYPE_CHECKING

import aiofiles

from gobbler_core.utils.file_handler import get_file_extension, validate_input_path
from gobbler_core.utils.frontmatter import count_words, create_document_frontmatter
from gobbler_core.utils.http_client import RetryableHTTPClient

if TYPE_CHECKING:
    from gobbler_core.providers.document import DocumentProvider

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = (".pdf", ".docx", ".pptx", ".xlsx", ".xls")


def _fix_checkboxes(content: str) -> str:
    """Fix checkbox rendering from PDF form fields.

    PDF form fields often render as awkward syntax like:
    - [ ] /Off (unchecked)
    - [ ] /Yes (checked)
    - [ ] /On (checked)

    This converts them to proper Unicode checkbox symbols:
    - ☐ (unchecked)
    - ☑ (checked)

    Args:
        content: Markdown content with raw checkbox syntax

    Returns:
        Content with fixed checkbox rendering
    """
    import re

    # Pattern: [ ] followed by /Off, /No, etc. = unchecked
    content = re.sub(r'\[ ?\] ?/Off\b', '☐', content)
    content = re.sub(r'\[ ?\] ?/No\b', '☐', content)

    # Pattern: [ ] followed by /Yes, /On, etc. = checked
    content = re.sub(r'\[ ?\] ?/Yes\b', '☑', content)
    content = re.sub(r'\[ ?\] ?/On\b', '☑', content)

    # Also handle standalone /Off and /Yes that might appear
    content = re.sub(r'\b/Off\b', '☐', content)
    content = re.sub(r'\b/Yes\b', '☑', content)
    content = re.sub(r'\b/On\b', '☑', content)

    return content


async def convert_document_to_markdown(
    file_path: str,
    enable_ocr: bool = True,
    service_url: str = "http://localhost:5001",
    metrics_callback: Callable[[str, int], None] | None = None,
    logger_instance: logging.Logger | None = None,
    provider: DocumentProvider | None = None,
) -> tuple[str, dict]:
    """Convert document to markdown using a document conversion provider.

    Uses a pluggable document conversion provider for the actual conversion.
    If no provider is specified, uses the default Docling Docker service.

    Args:
        file_path: Absolute path to document file
        enable_ocr: Enable OCR for scanned documents
        service_url: Docling service URL (default: "http://localhost:5001")
            Only used if provider is None
        metrics_callback: Optional callback for metrics tracking,
            called with (converter_type, size_bytes)
        logger_instance: Optional custom logger instance
        provider: Optional document conversion provider. If None, uses
            default DoclingProvider with the specified service_url.

    Returns:
        Tuple of (markdown_content, metadata)

    Raises:
        ValueError: Invalid file path or unsupported format
        RuntimeError: Service unavailable or conversion failed

    Example:
        # Using default provider
        markdown, metadata = await convert_document_to_markdown("document.pdf")

        # Using a specific provider
        from gobbler_core.providers.document import DoclingProvider
        provider = DoclingProvider(service_url="http://localhost:5001")
        markdown, metadata = await convert_document_to_markdown("document.pdf", provider=provider)
    """
    # Use custom logger if provided, otherwise use module logger
    log = logger_instance or logger

    # Validate file path
    error = validate_input_path(file_path, SUPPORTED_EXTENSIONS)
    if error:
        raise ValueError(error)

    file_format = get_file_extension(file_path)
    provider_name = provider.name if provider else "docling"

    log.info(
        "Starting document conversion",
        extra={
            "extra_fields": {
                "file_path": file_path,
                "file_format": file_format,
                "enable_ocr": enable_ocr,
                "provider": provider_name,
            }
        },
    )
    start_time = time.time()

    # Use provider-based conversion if a provider is specified
    if provider is not None:
        result = await provider.convert(Path(file_path), ocr=enable_ocr)
        markdown_content = result.markdown
        pages = result.pages
        word_count = result.word_count
    else:
        # Legacy path: use direct Docling HTTP calls
        markdown_content, pages, word_count = await _convert_with_docling(
            file_path, enable_ocr, service_url, log
        )

    # Post-process: Fix checkbox rendering from PDF form fields
    markdown_content = _fix_checkboxes(markdown_content)

    conversion_time_ms = int((time.time() - start_time) * 1000)

    # Create frontmatter
    frontmatter = create_document_frontmatter(
        file_path=file_path,
        doc_format=file_format,
        pages=pages,
        word_count=word_count,
        conversion_time_ms=conversion_time_ms,
    )

    # Combine frontmatter and markdown
    full_markdown = frontmatter + markdown_content

    # Track conversion size via callback if provided
    if metrics_callback is not None:
        metrics_callback("document", len(full_markdown))

    # Prepare metadata response
    metadata = {
        "file_path": file_path,
        "format": file_format,
        "pages": pages,
        "word_count": word_count,
        "conversion_time_ms": conversion_time_ms,
        "provider": provider_name,
    }

    log.info(
        "Document conversion completed",
        extra={
            "extra_fields": {
                "word_count": word_count,
                "pages": pages,
                "file_format": file_format,
                "provider": provider_name,
            }
        },
    )

    return full_markdown, metadata


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """Convert legacy .xls file to .xlsx format.

    Args:
        xls_path: Path to the .xls file

    Returns:
        Path to temporary .xlsx file (caller must delete)

    Raises:
        RuntimeError: If conversion fails
    """
    import os
    import tempfile

    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError as e:
        msg = (
            "xlrd and openpyxl are required for .xls file support. "
            "Install with: pip install xlrd openpyxl"
        )
        raise RuntimeError(msg) from e

    try:
        xls_book = xlrd.open_workbook(str(xls_path))
        xlsx_book = Workbook()

        # Remove the default sheet
        default_sheet = xlsx_book.active
        if default_sheet is not None:
            xlsx_book.remove(default_sheet)

        # Copy each sheet
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        # Write to temporary file
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(temp_fd)
        xlsx_book.save(temp_path)

        logger.debug("Converted %s to temporary %s", xls_path, temp_path)
        return Path(temp_path)

    except Exception as e:
        msg = f"Failed to convert .xls to .xlsx: {e}"
        raise RuntimeError(msg) from e


def _process_docling_response(result: dict) -> tuple[str, int, int]:
    """Process Docling API response and extract markdown content.

    Args:
        result: JSON response from Docling API

    Returns:
        Tuple of (markdown_content, pages, word_count)

    Raises:
        RuntimeError: If response indicates failure or has no content
    """
    if result.get("status") == "failure":
        errors = result.get("errors", ["Unknown error"])
        msg = f"Document conversion failed: {'; '.join(errors)}"
        raise RuntimeError(msg)

    if result.get("status") == "skipped":
        msg = (
            "Document conversion was skipped. The file may be corrupted or "
            "use an unsupported format variation."
        )
        raise RuntimeError(msg)

    document_data = result.get("document", {})
    markdown_content = document_data.get("md_content", "")

    if not markdown_content:
        msg = (
            "Failed to extract markdown from document. The document may be "
            "corrupted or password-protected."
        )
        raise RuntimeError(msg)

    word_count = count_words(markdown_content)
    pages = result.get("pages", 0) if "pages" in result else max(1, word_count // 300)

    return markdown_content, pages, word_count


async def _convert_with_docling(
    file_path: str,
    enable_ocr: bool,
    service_url: str,
    log: logging.Logger,
) -> tuple[str, int, int]:
    """Convert document using direct Docling HTTP calls (legacy path).

    Args:
        file_path: Path to document file
        enable_ocr: Enable OCR for scanned documents
        service_url: Docling service URL
        log: Logger instance

    Returns:
        Tuple of (markdown_content, pages, word_count)

    Raises:
        RuntimeError: If conversion fails
    """
    path = Path(file_path)
    temp_xlsx_path: Path | None = None

    # Handle legacy .xls files by converting to .xlsx first
    if path.suffix.lower() == ".xls":
        log.info("Converting legacy .xls file to .xlsx for processing")
        temp_xlsx_path = _convert_xls_to_xlsx(path)
        file_path = str(temp_xlsx_path)

    try:
        # Read file asynchronously
        try:
            async with aiofiles.open(file_path, "rb") as f:
                file_data = await f.read()
        except Exception as e:
            msg = f"Failed to read document file: {e}"
            raise RuntimeError(msg) from e

        filename = Path(file_path).name

        try:
            async with RetryableHTTPClient(timeout=120.0) as client:
                files = {"files": (filename, file_data)}
                data = {"to_formats": "md", "do_ocr": str(enable_ocr).lower()}

                response = await client.post(
                    f"{service_url}/v1/convert/file", files=files, data=data
                )
                response.raise_for_status()
                result = response.json()

        except Exception as e:
            error_str = str(e)
            error_type = type(e).__name__

            if "ConnectError" in error_type or "Connection" in error_str:
                msg = (
                    "Docling service unavailable. The service may not be running. "
                    "Start with: docker-compose up -d docling"
                )
                raise RuntimeError(msg) from e
            msg = f"Document conversion failed: {e}"
            raise RuntimeError(msg) from e

        return _process_docling_response(result)

    finally:
        # Clean up temporary .xlsx file if we created one
        if temp_xlsx_path is not None and temp_xlsx_path.exists():
            try:
                temp_xlsx_path.unlink()
                log.debug("Cleaned up temporary file: %s", temp_xlsx_path)
            except OSError as e:
                log.warning("Failed to clean up temporary file %s: %s", temp_xlsx_path, e)
